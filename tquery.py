from sys import argv
import pymysql
import importlib
import csv
import sys
import re
import glob
import os
import pandas as pd
from pandas import ExcelWriter
import glob
import argparse
import random
import string
from os.path import abspath, dirname
from arg_parser import parse_argv  # Import the parse_argv function from the external module
from query import openquery
import json
import time
import warnings
import subprocess
from Database import *
import openpyxl
from openpyxl import load_workbook

warnings.filterwarnings("ignore")

# to make the current dir where the script resides
current_wrk_dir = os.getcwd()
abspath = os.path.abspath(__file__)
dname = os.path.dirname(abspath)
os.chdir(dname)


def convert_dataframe_to_json(dataframe):
    json_data = dataframe.to_json(orient='records')
    json_body = json.loads(json_data)
    json_formatted_str = json.dumps(json_body, indent=4)
    return json_formatted_str

# Define a function to dynamically import a module by name
def import_module(module_name):
    try:
        imported_module = importlib.import_module(module_name)
        print(f"Module '{module_name}' imported successfully.")
        return imported_module
    except ImportError:
        print(f"Failed to import module '{module_name}'.")
        return None

# Function to fetch data from a CSV file and return both the DataFrame and the CSV filename
def fetch_data_from_csv(csv_filename):
    try:
        # Read the CSV file using pandas
        df = pd.read_csv(csv_filename)
        return df, csv_filename
    except Exception as e:
        print(f"Error reading CSV file: {str(e)}")
        return None
# Function to fetch data from an Excel file and return a dictionary of DataFrames
def fetch_data_from_excel(excel_filename):
    try:
        data_dict = {}  # Initialize data_dict as an empty dictionary
        # Load the Excel file using openpyxl
        wb = load_workbook(excel_filename)
        for sheet_name in wb.sheetnames:
            df = pd.read_excel(excel_filename, sheet_name=sheet_name)
            data_dict[sheet_name] = df
        return data_dict
    except Exception as e:
        print(f"Error reading Excel file: {str(e)}")
        return None

# Parse command-line arguments using the imported function
myargs, config_path, csv_module_name, excel_filename, query, out_path = parse_argv(sys.argv[1:])
print(myargs)

if __name__ == '__main__':
    queriesResult = []
    finalSheetNames = []
    block = ""
    # json_data_dict = {}
    start_time = time.time()
    conn = get_database_connection(config_path)

    # Check for the 'excelfile' argument
    if 'excelfile' in myargs:
        excel_filenames = myargs['excelfile']
        excel_files = excel_filenames.split(':')
        excel_data_dict = {}

        for excel_file in excel_files:
            data_from_excel = fetch_data_from_excel(excel_file)
            if data_from_excel is not None:
                excel_data_dict[excel_file] = data_from_excel
                print(f"Data from Excel file {excel_file}:")
                for sheet_name, df in data_from_excel.items():
                    print(f"Sheet: {sheet_name}")
                    print(df)

    # Check for the 'csvfile' argument
    if 'csvfile' in myargs:
        csv_filenames = myargs['csvfile']  # Get the provided CSV filenames as a list
        csv_files = csv_filenames.split(':')  # Split the filenames by ':' into a list
        # Created a dictionary to store DataFrames with CSV filenames as keys
        data_dict = {}

        for csv_file in csv_files:
            data_from_csv, csv_name = fetch_data_from_csv(csv_file)
            print(data_from_csv)
            if data_from_csv is not None:
                data_dict[csv_name] = data_from_csv  # Store the DataFrame with the CSV filename as the key
                # print(data_dict)

    # Check if csv is in myargs
    if 'csv' in myargs:
        csv_module_name = myargs['csv']  # Get the provided CSV module name
        if csv_module_name:
            try:
                imported_module_csv = importlib.import_module(csv_module_name)
                print(f"Module '{csv_module_name}' imported successfully.")

                # Check if the imported module has a 'main' function
                if hasattr(imported_module_csv, "main") and callable(imported_module_csv.main):
                    imported_module_csv.main("Input_vectors.csv", "mapping.csv")
                else:
                    print(f"Module '{csv_module_name}' does not have a 'main' function.")
            except ImportError:
                print(f"Failed to import module '{csv_module_name}' for CSV data")

    replacedData = openquery(myargs)
    if replacedData is None:
        pass
    else:
        if conn:
            print("connected successfully")
        else:
            print("Not Connected")
        splitString = replacedData.split(";")
        try:
            cursor = conn.cursor()
            success = True  # Flag for the success of all operations

            for cmdoneonly in splitString:
                cmdoneonly = cmdoneonly.replace('\n', ' ').replace('\n\n', '').lower().strip()

                if cmdoneonly.strip():
                    if cmdoneonly.startswith(('update', 'delete', 'insert')):
                        try:
                            cursor.execute(cmdoneonly)
                            conn.commit()  # Commit the transaction if the operation is successful
                            print("Operation completed successfully")
                        except Exception as e:
                            conn.rollback()  # Rollback the transaction if an error occurs
                            print(f"Operation failed: {str(e)}")
                            success = False  # Set the success flag to False
                    else:
                        cursor.execute(cmdoneonly)
                        results = cursor.fetchall()
                        if results:
                            df = pd.DataFrame(results, columns=[x[0] for x in cursor.description])
                            queriesResult.append(df)
                            # check if json is in myargs
                            if 'json' in myargs:
                                json_formatted_str = convert_dataframe_to_json(df)
                                print("json output is:", json_formatted_str)

                            else:
                                # Dynamically import the module for JSON data
                                json_module_name = myargs.get('json_module')
                                if json_module_name is None:
                                    print("Please provide a Input python file name which will be used as Json module")
                                else:
                                    imported_module_json = import_module(json_module_name)
                                # imported_module_json = import_module(json_module_name)

                                if imported_module_json:
                                    if hasattr(imported_module_json, "main") and callable(
                                            imported_module_json.main):
                                        result = imported_module_json.main(myargs, queriesResult)
                                        print(result)  # Print or return the result as needed
                                    else:
                                        print(
                                            f"Module '{json_module_name}' for JSON data does not have a 'main' function.")

                                sheetName = re.compile(r'^select\s+.+from\s+([a-z_]+)')
                                for names in sheetName.finditer(cmdoneonly):
                                    finalSheetNames.append(names.group(1))
                                folderName = myargs['query'][:-6]
                                query_file = os.path.basename(myargs['query'])  # Get the query filename
                                data = myargs.get('out_path', current_wrk_dir)
                                # Check if there are values in double quotes
                                double_quoted_values = re.findall(r'"([^"]*)"', cmdoneonly)

                                # Use folderName as the top-level folder if there are values in double quotes, else use the current working directory
                                base_folder = os.path.join(current_wrk_dir,
                                                           folderName) if double_quoted_values else current_wrk_dir

                                # Create a top-level folder for each value in double quotes or use folderName
                                for value in double_quoted_values:
                                    # Combine value and query_file to create folder structure
                                    folder_path = os.path.join(base_folder, value, query_file)
                                    # Ensure the folder structure exists
                                    path = os.path.join(base_folder, folder_path)
                                    os.makedirs(path, exist_ok=True)

                                    # Save the DataFrame to a CSV file inside the folder_path
                                    for n, df in enumerate(queriesResult):
                                        csv_filename = os.path.join(path, f'{finalSheetNames[n]}.csv')
                                        df.to_csv(csv_filename, index=False)
                                    print("CSV files created successfully in", csv_filename)

                                # Save the DataFrame to a CSV file in the base_folder if there are no values in double quotes
                                if not double_quoted_values:
                                    for n, df in enumerate(queriesResult):
                                        csv_filename = os.path.join(base_folder, f'{finalSheetNames[n]}.csv')
                                        df.to_csv(csv_filename, index=False)
                                    print("CSV files created successfully in", csv_filename)
                        else:
                            print("The query result doesn't have any data to showcase")
            cursor.close()

            if success:
                print("All operations completed successfully")
            else:
                print("Some operations failed, transaction rolled back")
            # print("script execution completed")
        except pymysql.err.ProgrammingError as except_detail:
            print("pymysql.err.ProgrammingError: «{}»".format(except_detail))
        finally:
            conn.close()
            end_time = time.time()
            print("Script successfully completed in:", end_time - start_time, "seconds")
